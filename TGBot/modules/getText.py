from openpyxl import load_workbook


class getText:

    @staticmethod
    def maxrowSearch(ws):
        if ws.max_row == 1:
            if ws.cell(row=1, column=1).value == "" or ws.cell(row=1, column=1).value == None:
                return 1
            else:
                return 2
        else:
            return int(ws.max_row + 1)

    @staticmethod
    def Addindex(man):
        if man == 'admin':
            return 1
        elif man == 'pupil':
            return 2
        elif man == 'noname':
            return 3
        elif man == 'other':
            return 4

    def __init__(self, man):
        self.path = "./Data/Texts.xlsx"
        self.wb = load_workbook(self.path)

        self.index = self.Addindex(man)

        self.sheets = {
            1: 'Admin',
            2: 'Pupil',
            3: 'Noname',
            4: 'Other'
        }

    def gettext(self, param):
        """
        Достает текст исходя из параметра
        """
        ws = self.wb.get_sheet_by_name(self.sheets[self.index])
        for i in range(2, self.maxrowSearch(ws)):
            if param == ws.cell(row=i, column=1).value:
                if ws.cell(row=i, column=2).value != "" and ws.cell(row=i, column=2).value != None:
                    return ws.cell(row=i, column=2).value
                else:
                    return param